import os
import openpyxl

file_path = r"c:\Users\longn\rosesalon\Customer-monitoring.xlsx"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    exit(1)

try:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    print("Workbook sheet names:", wb.sheetnames)
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        rows = list(sheet.iter_rows(max_row=20, values_only=True))
        for i, row in enumerate(rows):
            print(f"Row {i+1}: {row}")
            
except Exception as e:
    print("Error reading with openpyxl:", e)
    
    # Try pandas if openpyxl fails or to see if we have it
    try:
        import pandas as pd
        df = pd.read_excel(file_path)
        print("Pandas read head:")
        print(df.head(20))
    except Exception as pe:
        print("Error reading with pandas:", pe)
