import openpyxl

wb = openpyxl.load_workbook('ROSEBEAU-DAILY-SALES.xlsx', data_only=False)
sheet = wb['2025 NEW DSR']

print(f"{'Row':<5} | {'Col P (Val/Formula)':<25} | {'Col Q':<15} | {'Col R':<15}")
print("-" * 65)

for row in range(16, 33):
    val_p = sheet.cell(row=row, column=16).value
    val_q = sheet.cell(row=row, column=17).value
    val_r = sheet.cell(row=row, column=18).value
    print(f"{row:<5} | {str(val_p):<25} | {str(val_q):<15} | {str(val_r):<15}")
