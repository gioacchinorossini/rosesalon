import os
import openpyxl

file_path = r"c:\Users\longn\rosesalon\Customer-monitoring.xlsx"

if not os.path.exists(file_path):
    print("File not found.")
    exit(1)

wb = openpyxl.load_workbook(file_path, read_only=True)
sheet = wb.active
print("Total rows:", sheet.max_row)

non_empty_rows = []
for idx, row in enumerate(sheet.iter_rows(values_only=True)):
    if any(cell is not None for cell in row):
        non_empty_rows.append((idx + 1, row))

print(f"Found {len(non_empty_rows)} non-empty rows:")
for r_idx, r_val in non_empty_rows:
    print(f"Row {r_idx}: {r_val}")
