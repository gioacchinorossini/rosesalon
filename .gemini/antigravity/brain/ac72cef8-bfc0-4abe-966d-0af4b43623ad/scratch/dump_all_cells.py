import openpyxl

wb = openpyxl.load_workbook('ROSEBEAU-DAILY-SALES.xlsx', data_only=False)
sheet = wb['2025 NEW DSR']

non_empty = []
for r in range(1, sheet.max_row + 1):
    for c in range(1, sheet.max_column + 1):
        val = sheet.value = sheet.cell(row=r, column=c).value
        if val is not None:
            non_empty.append((r, c, val))

for r, c, val in non_empty:
    coord = f"{openpyxl.utils.get_column_letter(c)}{r}"
    print(f"Cell {coord:<5}: {val}")
